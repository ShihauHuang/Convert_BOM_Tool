from openpyxl import Workbook, load_workbook, __version__
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from Logger import Create_Logger, logging
from re import search
from datetime import datetime
from configparser import ConfigParser
import tkinter as tk
from tkinter import StringVar, ttk
from tkinter.filedialog import askopenfilenames
from os import mkdir, path

# 當前的 openpyxl 版本為 3.2.0b1，非正式 release 直接從 github 上下載
# https://openpyxl.readthedocs.io/en/latest/index.html openpyxl 文件

def Get_Original_Data( xlsx_path ) :

    logging.debug('xlsx_path : ' + xlsx_path)
    wb = load_workbook ( xlsx_path ) # 讀取指定 xlsx 檔案
    #original_sheet = wb['原稿']
    original_sheet = wb.active
    #BOM_sheet = openpyxl.Workbook()

    product_dict = {}
    kit_list = [] # 此為用來排序 BOM 的零件項目 (依照英數字排序)
    serial = -1
    for row in original_sheet.iter_rows ( min_row = int(conf.get('Get_Original_Data_parameters' , 'start_row_num')) , max_col = int(conf.get('Get_Original_Data_parameters' , 'total_col_amount')) ) : # 以列（row）方向逐一疊代處理
    
        if row[0].value != None : # 判斷 等級 0 是否為空值

            serial = serial + 1 #　起始為 -1 這樣就會變成 0
            product_dict[serial] = {}

            # Example :　A503-01-鋼製中央實驗桌.藥架.壁櫃2550*1500*825/2400
            product_id = row[0].value.rsplit('-' , 1)[0]  # 從後面分割 - 1次
            product = row[0].value.rsplit('-' , 1)[1]

            product_name = search( '(.+?)\d' , product).group(1) # 分析產品名稱
            product_dict[serial]['product_name'] = product_name
            
            product_size = search( '(\d.+)' , product).group(1) # 分析產品尺寸
            product_dict[serial]['product_size'] = product_size

            product_amount = row[2].value # 分析產品數量
            product_dict[serial]['product_amount'] = int(product_amount)

            product_dict[serial]['product_id'] = product_id

            product_dict[serial]['product_kits'] = {}

        else : # 等級 0 為空的話，代表目前為其底下的零件
            kit_full = row[1].value 
            #kit_name = kit_full.rsplit('-' , 1)[1]
            kit_name = search( '\d{4}-?(.+)'  , kit_full ).group(1) # (BB-6002-鐵腳695x825x533mm(庫存)) 判斷依據 4 數字後不管有無出現 "-" 的後方所有
            kit_amount = row[2].value
            
            #product_dict[product_name]['product_kits'][kit_name] = int(kit_amount) * product_dict[product_name]['product_amount'] # 實際零件數量要乘上產品數量
            product_dict[serial]['product_kits'][kit_name] = int(kit_amount)

            kit_list.append(kit_full) # 將原始加入 kit_list 做英數字排序用

    logging.debug('product_dict : ' + str(product_dict))
    
    logging.debug('original_kit_list : ' + str(kit_list))

    kit_list.sort() # 按照順序排列
    # 由後比對是否有出現過一樣的
    for i in range ( len(kit_list)-1 , 0 , -1 ) : 
        now_item = search( '\d{4}-?(.+)'  , kit_list[i] ).group(1) 
        for j in range ( i-1 , -1 , -1 ) :
            previous_item = search( '\d{4}-?(.+)'  , kit_list[j] ).group(1)
            if previous_item == now_item :
                del kit_list[i]
                break
    logging.debug('kit_list : ' + str(kit_list))

    final_kit_list = []
    for a in kit_list :
        final_kit_list.append(search( '\d{4}-?(.+)'  , a ).group(1))

    logging.debug('final_kit_list : ' + str(final_kit_list))
    return product_dict , final_kit_list

def Auto_set_column_width ( col_cells ) :

    logging.debug('col_cells : ' + str(col_cells))
    gold_rate = float(conf.get('BOM_Format' , 'col_width_rate_for_ABC_col')) # 1.61803
    tmp_list = []
    for cell in col_cells : 
        value = cell[0].value #(<Cell 'BOM表'.A5>,) 需要用 cell[0]
        logging.debug ('value : ' + value)
        logging.debug ('value len : ' + str(len(value.encode('big5'))))
        tmp_list.append( len(value.encode('big5')) )
    logging.debug('tmp_list : ' + str(tmp_list))

    max_width = max(tmp_list) * gold_rate
    logging.debug('max_width : ' + str(max_width))

    return max_width
        
def Create_BOM (product_dict , kit_list , file_name , folder_full_path, _Date, file_time) :
    
    logging.info('Current file : ' + str(file_name))
    wb = Workbook() # 建立新的 excel
    sheet = wb.active # 取得當前 sheet
    sheet.title = conf.get('BOM_Format' , 'sheet_name')

    total_row = 4 + len(product_dict) # 預設 4 列為固定
    total_col = 5 + len(kit_list) # 預設 5 欄為固定
    logging.debug( 'total_row : ' + str(total_row))
    logging.debug( 'total_col : ' + str(total_col))

    font_family = conf.get('BOM_Format' , 'font_family') # 預設為 新細明體
    font_size = conf.get('BOM_Format' , 'font_size') # 預設為 16
    medium_font_size = conf.get('BOM_Format' , 'medium_font_size') # 預設為 28
    large_font_size = conf.get('BOM_Format' , 'large_font_size') # 預設為 40

    # 固定部分 ========================================================================================
    # merge 之後只有第一格能改寫數值，其餘屬性為 read-only
    # 先給 Border 在 merge
    sheet['A1'].border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet.merge_cells( start_row=1, start_column=1, end_row=1, end_column=3) 
    sheet['A1'].value = file_name + ' BOM表'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet['A1'].font = Font(name=font_family, size=font_size, bold=False, color='FF0000')
    
    sheet['D1'].border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet.merge_cells( start_row=1, start_column=4, end_row=1, end_column=total_col)
    sheet['D1'].value = conf.get('BOM_Format' , 'D1_default_value')
    sheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet['D1'].font = Font(name=font_family, size=font_size, bold=False, color='FF0000')

    sheet['A2'].border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet.merge_cells( start_row=2, start_column=1, end_row=2, end_column=3) 
    sheet['A2'].value = conf.get('BOM_Format' , 'A2_default_value')
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet['A2'].font = Font(name=font_family, size=font_size, bold=True, color='FF0000')

    sheet['D2'].border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet.merge_cells( start_row=2, start_column=4, end_row=2, end_column=total_col)
    sheet['D2'].value = conf.get('BOM_Format' , 'D2_default_value')
    sheet['D2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet['D2'].font = Font(name=font_family, size=font_size, bold=False, color='FF0000')

    sheet['A3'].border = Border( right=Side(style='thin') ) 
    sheet['A3'].value = '圖\n號'
    sheet['A3'].font = Font(name=font_family, size=font_size, bold=False)
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  

    size_default = InlineFont(sz=font_size, color='FF0000')
    
    sheet['B3'].border = Border( right=Side(style='thin') ) 
    #sheet['B3'].value = '品  名'
    #sheet['B3'].font = Font(name=font_family, size=font_size, bold=False)
    size_medium = InlineFont(sz=medium_font_size, color='FF0000', b=True)
    B3_specify = conf.get('BOM_Format' , 'B3_default_value')
    sheet["B3"] = CellRichText([TextBlock(font=size_default, text='\n'*8 + "品  名" + '\n'*7), TextBlock(font=size_medium, text=B3_specify)])
    sheet['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  

    sheet['C3'].border = Border( right=Side(style='medium') ) 
    size_large = InlineFont(sz=large_font_size, color='FF0000', b=True)
    sheet["C3"] = CellRichText([TextBlock(font=size_default, text="尺寸\nmm\n"), TextBlock(font=size_large, text="交期\n" + _Date.strip())])
    sheet['C3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet['D3'].border = Border( right=Side(style='thin') ) 
    sheet['D3'].value = '數\n量'
    sheet['D3'].font = Font(name=font_family, size=font_size, bold=False)
    sheet['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  

    sheet['E3'].border = Border( right=Side(style='medium') ) 
    sheet['E3'].value = '單\n位'
    sheet['E3'].font = Font(name=font_family, size=font_size, bold=False)
    sheet['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  

    sheet['A4'].border = Border( left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet.merge_cells( start_row=4, start_column=1, end_row=4, end_column=3) 
    sheet['A4'].value = '總計數量'
    sheet['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet['A4'].font = Font(name=font_family, size=font_size, bold=False)

    sheet['D4'].border = Border( left=Side(style='medium'), right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium') ) 
    sheet['E4'].border = Border( left=Side(style='thin'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') ) 

    for row in sheet.iter_rows ( min_row=4, max_row= 4, min_col=6, max_col= total_col ) : # 對第四列寫入固定公式
        for _cell in row :
            _cell.border = Border( right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='medium') ) 
            _cell.value = '=SUM(INDIRECT("R1000C:R[+1]C",FALSE))'
            _cell.font = Font(name=font_family, size=font_size, bold=False)
            _cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    sheet.cell(row=4 , column=total_col ).border = Border( right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium') )  # 把最右邊的粗框弄回去

    # 零件列表 =======================================================================================
    for row in sheet.iter_rows ( min_row=3, max_row= 3, min_col=6, max_col= total_col ) : # 對第三列做動作
        for i , _cell in enumerate(row) :

            # 此為將文字轉為直行，並將裡面部分情況做特別判斷
            # Example : 靠邊立柱825/2400mm/深70mm桌面線槽型
            convert_content_to_vertical = ''
            digit_flag = False # 用於紀錄連續數字
            mm_flag = False # 用於紀錄連續 mm
            for _char in kit_list[i] :
                if  _char.isdigit() : # 是數字需連續
                    if mm_flag : # 因程式執行先後順序問題，如果 mm 後面接數字需特別判斷
                        convert_content_to_vertical = convert_content_to_vertical + '\n'
                        mm_flag = False
                    convert_content_to_vertical = convert_content_to_vertical + _char
                    digit_flag = True
                    continue
                elif digit_flag : # 代表連續性的數字結束，需要給一個換行
                    convert_content_to_vertical = convert_content_to_vertical + '\n'
                    digit_flag = False

                if  _char == 'm' : # 是 m 需連續
                    convert_content_to_vertical = convert_content_to_vertical + _char
                    mm_flag = True
                    continue
                elif mm_flag : # 代表連續性的 m 結束，需要給一個換行
                    convert_content_to_vertical = convert_content_to_vertical + '\n'
                    mm_flag = False

                if _char == '(' :
                    convert_content_to_vertical = convert_content_to_vertical + '︵\n'

                elif _char == ')' :
                    convert_content_to_vertical = convert_content_to_vertical + '︶\n'

                elif _char == 'x' :
                    convert_content_to_vertical = convert_content_to_vertical + '*\n'

                else :
                    convert_content_to_vertical = convert_content_to_vertical + _char + '\n'
            _cell.border = Border( right=Side(style='thin')) 
            _cell.value = convert_content_to_vertical.strip()
            _cell.font = Font(name=font_family, size=font_size, bold=False)
            _cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)  
    sheet.cell(row=3 , column=total_col ).border = Border( right=Side(style='medium'))  # 把最右邊的粗框

    # 產品欄位 =======================================================================================
    data_row = sheet.iter_rows ( min_row=5, max_row= total_row, min_col=1, max_col= total_col )  # 寫入產品及其零件數量，從 A5 開始
    for index, row in enumerate(data_row) :
        row[0].value = product_dict[index]['product_id']
        row[1].value = product_dict[index]['product_name']
        row[2].value = product_dict[index]['product_size']
        row[3].value = product_dict[index]['product_amount']
        row[4].value = '台'
        for i, kit in enumerate(kit_list) : # 跑 F3:3 迴圈零件名稱
            if kit in product_dict[index]['product_kits']: # 比對零件
                sheet.cell(row= index + 5, column= i + 6).value = '=' + str(product_dict[index]['product_kits'][kit]) + '*D' + str(index + 5)
        
        for j, _cell in enumerate(row) : 

             # 依照原本提供表格加粗框
            if index == len(product_dict) - 1 : # 如果是最後一列，則底部加粗框
                if j == 3 : # 特定位置加粗框
                    _cell.border = Border( left=Side(style='medium'),  right=Side(style='thin'), bottom=Side(style='medium') ) 
                elif j == 4 : # 特定位置加粗框
                    _cell.border = Border( left=Side(style='thin'),  right=Side(style='medium'), bottom=Side(style='medium') ) 
                elif j == total_col - 1 : # 最右邊
                    _cell.border = Border( left=Side(style='thin'),  right=Side(style='medium'), bottom=Side(style='medium') ) 
                else :
                    _cell.border = Border( right=Side(style='thin'), bottom=Side(style='medium') ) 
            else :
                if j == 3 :
                    _cell.border = Border( left=Side(style='medium'),  right=Side(style='thin'), bottom=Side(style='thin') ) 
                elif j == 4 :
                    _cell.border = Border( left=Side(style='thin'),  right=Side(style='medium'), bottom=Side(style='thin') ) 
                elif j == total_col - 1 : # 最右邊
                    _cell.border = Border( left=Side(style='thin'),  right=Side(style='medium'), bottom=Side(style='thin') ) 
                else :
                    _cell.border = Border( right=Side(style='thin'), bottom=Side(style='thin') ) 

            _cell.font = Font(name=font_family, size=font_size, bold=False)
            if j == 1 :
                _cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  
            else :
                _cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    
    col_A_width = Auto_set_column_width ( sheet['A5':'A' + str(total_row)] )
    if col_A_width < 10 : col_A_width = 10
    col_B_width = Auto_set_column_width ( sheet['B5':'B' + str(total_row)] )
    if col_B_width < 38 : col_B_width = 38
    col_C_width = Auto_set_column_width ( sheet['C5':'C' + str(total_row)] )
    if col_C_width < 38 : col_C_width = 31

    # 設定欄位寬度 ====================================================================================
    sheet.column_dimensions["A"].width = col_A_width
    sheet.column_dimensions["B"].width = col_B_width
    sheet.column_dimensions["C"].width = col_C_width
    for col_w in range(4 , total_col + 1 ) : 
        sheet.column_dimensions[get_column_letter(col_w)].width = float(conf.get('BOM_Format' , 'col_D_to_end_width')) # 7
    
    # 設定列位高度 ====================================================================================
    sheet.row_dimensions[1].height = float(conf.get('BOM_Format' , 'row_1_height')) # 25
    sheet.row_dimensions[2].height = float(conf.get('BOM_Format' , 'row_2_height')) # 25
    sheet.row_dimensions[3].height = float(conf.get('BOM_Format' , 'row_3_height')) # 410
    sheet.row_dimensions[4].height = float(conf.get('BOM_Format' , 'row_4_height')) # 40
    for row_h in range(5 , total_row + 1 ) : 
        sheet.row_dimensions[row_h].height = float(conf.get('BOM_Format' , 'row_5_to_end_height')) # 26

    # 設定凍結 ========================================================================================
    freeze_col_check = conf.get('BOM_Format' , 'freeze_col_check') # 預設為 disable
    freeze_row_check = conf.get('BOM_Format' , 'freeze_row_check') # 預設為 disable
    freeze_col = conf.get('BOM_Format' , 'freeze_col') # 預設為 F
    freeze_row = conf.get('BOM_Format' , 'freeze_row') # 預設為 4
    if freeze_row_check != 'disable' and  freeze_col_check != 'disable':
        sheet.freeze_panes = freeze_col + freeze_row
    elif freeze_row_check != 'disable' and  freeze_col_check == 'disable':
        sheet.freeze_panes = 'A' + freeze_row # 使用 A 等於欄沒有凍結，此函數欄列都必須給予
    elif freeze_row_check == 'disable' and  freeze_col_check != 'disable':
        sheet.freeze_panes = freeze_col + '1' # 使用 1 等於列沒有凍結，此函數欄列都必須給予

    try :
        wb.save(folder_full_path + '/' + file_name +  '-' + file_time + '.xlsx')
        logging.info('Convert "' + file_name + '" successfully.')
        return True
    except Exception as e :
        logging.warning('Something fail : ' + str(e))
        return False

def Create_UI ( ):
    app = tk.Tk()
    app.title('Excel 轉換工具 v' + tool_version)
    app.resizable (False , False) # 不可縮放
    app.geometry('500x345+368+207') # 設定初始值 寬x高+X座標+Y座標 (座標不一定要寫入)

    # 瀏覽按鈕 =========================================================================== 
    Browse_btn = tk.Button(app , text = '選擇 Excel 檔' , font=("Calibri 20"), command = lambda : Browse_files_and_run(
        A2_C2_Entry.get().strip(), 
        D1_1_Entry.get().strip(), 
        D2_2_Entry.get().strip(), 
        B3_Entry.get().strip(), 
        Date_Entry.get().strip(), 
        Progress_Bar, 
        Percent_Label,
        )
    )
    Browse_btn.place(x=10 , y=185 , width=480 , height=70)
    Browse_btn.place(x=10 , y=230 , width=480 , height=70)

    # 基本預設 ===========================================================================
    A2_C2_Label = tk.Label(app , text = 'A2:C2' , font=("Calibri 14"))
    A2_C2_Label.place(x=3 , y=10 , width=60 , height=35)
    A2_C2_string = StringVar()
    A2_C2_Entry = tk.Entry(app  , font=("Calibri 14"), textvariable = A2_C2_string )
    A2_C2_value = conf.get('BOM_Format' , 'A2_default_value').strip()
    A2_C2_string.set(A2_C2_value)
    A2_C2_Entry.place(x=65 , y=10 , width=420 , height=35)    
    
    D1_1_Label = tk.Label(app , text = 'D1:1' , font=("Calibri 14"))
    D1_1_Label.place(x=3 , y=55 , width=60 , height=35)
    D1_1_string = StringVar()
    D1_1_Entry = tk.Entry(app  , font=("Calibri 14"), textvariable = D1_1_string )
    D1_1_value = conf.get('BOM_Format' , 'D1_default_value').strip()
    D1_1_string.set(D1_1_value)
    D1_1_Entry.place(x=65 , y=55 , width=420 , height=35)

    D2_2_Label = tk.Label(app , text = 'D2:2' , font=("Calibri 14"))
    D2_2_Label.place(x=3 , y=100 , width=60 , height=35)
    D2_2_string = StringVar()
    D2_2_Entry = tk.Entry(app  , font=("Calibri 14"), textvariable = D2_2_string )
    D2_2_value = conf.get('BOM_Format' , 'D2_default_value').strip()
    D2_2_string.set(D2_2_value)
    D2_2_Entry.place(x=65 , y=100 , width=420 , height=35)

    B3_Label = tk.Label(app , text = 'B3' , font=("Calibri 14"))
    B3_Label.place(x=3 , y=145 , width=60 , height=35)
    B3_string = StringVar()
    B3_Entry = tk.Entry(app  , font=("Calibri 14"), textvariable = B3_string )
    B3_value = conf.get('BOM_Format' , 'b3_default_value').strip()
    B3_string.set(B3_value)
    B3_Entry.place(x=65 , y=145 , width=420 , height=35)

    Date_Label = tk.Label(app , text = '交期' , font=("Calibri 14"))
    Date_Label.place(x=3 , y=190 , width=60 , height=35)
    Date_Entry = tk.Entry(app  , font=("Calibri 14") )
    Date_Entry.place(x=65 , y=190 , width=420 , height=35)
    Date_Entry.focus()

    # 進度條 =============================================================================
    Progress_Bar = ttk.Progressbar (app , orient = tk.HORIZONTAL, length = 100, mode = 'determinate') 
    Progress_Bar.place(x=10,y=305,width=450,height=35)
    Percent_Label = tk.Label(app, font=("Calibri 10") , text = "--%")
    Percent_Label.place(x=460,y=308,width=40,height=30)

    app.mainloop()

def Browse_files_and_run (A2_C2, D1_1, D2_2, B3, _Date, Progress_Bar, Percent_Label) :

    logging.debug('A2_C2 : ' + A2_C2)
    logging.debug('D1_1 : ' + D1_1)
    logging.debug('D2_2 : ' + D2_2)
    logging.debug('B3 : ' + B3)
    logging.debug('_Date : ' + _Date)

    file_list = askopenfilenames (title="選擇 excel 檔" , filetypes = [
        ('Excel' , '.xlsx'), 
        ('Excel' , '.xls'),
        ])
    if file_list != "" :

        progress_update (Progress_Bar , Percent_Label, 0 )

        # 寫入 cfg 預設值
        conf.set ('BOM_Format' , 'A2_default_value' , A2_C2)
        conf.set ('BOM_Format' , 'D1_default_value' , D1_1)
        conf.set ('BOM_Format' , 'D2_default_value' , D2_2)
        conf.set ('BOM_Format' , 'b3_default_value' , B3)
        conf.write(open('config.cfg', 'w', encoding='utf-8-sig'))

        folder_path = file_list[0].rsplit('/' , 1)[0] # 直接讀取第一個元素取得 path
        now = datetime.now()
        folder_time = now.strftime('%Y_%m_%d %H-%M-%S')
        file_time = now.strftime('%Y%m%d%H%M')
        folder_full_path = folder_path + '/' + folder_time + ' (Converted)'
        logging.debug('folder_full_path : ' + folder_full_path)
        if not path.exists(folder_full_path): # 建立資料夾存放轉換後的檔案
            mkdir(folder_full_path)

        interval = float (100 / len(file_list))
        for i, file_path in enumerate(file_list) :
            product_dict , kit_list = Get_Original_Data( file_path )
            file_name = file_path.rsplit('/' , 1)[1].rsplit('.')[0]
            if Create_BOM (product_dict , kit_list , file_name , folder_full_path, _Date, file_time) :
                progress_update (Progress_Bar , Percent_Label, (i+1)*interval )

def progress_update (Progress_Bar , Percent_Label, value) :
    Progress_Bar['value'] = value
    Progress_Bar.update_idletasks()
    Percent_Label['text'] = str(int(value)) + "%"
    Percent_Label.update_idletasks()

if __name__ == "__main__" :

    tool_version = '1.1'

    # 讀取 config 檔案
    conf = ConfigParser()
    conf.read('config.cfg' , encoding='utf-8-sig')

    Create_Logger()

    logging.debug(__version__)

    Create_UI()
